#  import required modules
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import sys
import time


# do I need to import os and use os.getcwd if the spreadsheet is in the same folder as main?

workbook = load_workbook("beyblade_db.xlsx")
Primary = workbook['Primary']
user1 = workbook['user1']
user2 = workbook['user2']

# use openpyxl.utils get_column_letter feature to convert column letters to numbers
def list_primary():
    for row in range(3, 6):
        for col in range(1, 5):
            char = get_column_letter(col)
            print(Primary[char + str(row)].value, end = " ")

# uses an alternate way to retrieve a range of rows
def list_user1():
    row=user1.max_row
    column = user1.max_column
    for i in range(3, row + 1):
        for j in range(1, column + 1):
            print(user1.cell(i, j).value, end = " ")

# another, simpler for loop to retrieve a range of rows
def list_user2():
    for i in range(1, 5):
        print(i, user2.cell(row=3, column=i).value)

# list menu options
def menu():
    print("\n[1] Views the master Beyblades list.")
    print("[2] Views User 1's Beyblades list.")
    print("[3] Views User 2's Beyblades list.")
    print("[0] Exit the program.")

# initialize menu
menu()

# introduce program and prompt for user input selection
def get_option_int():
    option = (input("""Welcome to the BeyBlade Organizing Interface program!
Please choose an option from above and enter the option number.
"""))
    # except error if user inputs a non-integer
    try:
        user_option = int(option)
        return user_option
    except ValueError:
        print("Only numbers coinciding with menu choices are accepted.")
        menu()
        return(get_option_int())

# initialize prompt for user input
option = get_option_int()

# while loop for menu
while option != 0:
    if option == 1:
        list_primary()
        #initiates first option
    elif option == 2:
        list_user1()
        # initiates second option
    elif option == 3:
        list_user2()
        # initiates third option
    else:
        print('Please choose an option from the menu:')
        #improper menu selection

    menu()
    option = get_option_int()

# Begin outro message
print("Good luck in your next BeyBattle!")

#Stylize the outro referencing BeyBlade gameplay
outro = "3... 2... 1... Let it RIP!"

for i in outro.split():
    sys.stdout.write("{} ".format(i))
    sys.stdout.flush()
    time.sleep(.4)

# Features addressed per Project Requirements:

# COMPLETED (menu)
# Implement a “master loop” console application where the user can repeatedly enter commands/perform
# actions, including choosing to exit the program

# Create a dictionary or list, populate it with several values, retrieve at least one value, and use it in your program
# PARTIALY COMPLETTED (via pyxl)
# Read data from an external file, such as text, JSON, CSV, etc and use that data in your application

# POSIBLY COMPLETED? (defined menu, each worksheet of the workbook, and function that verifies input and excepts error.)
# Create and call at least 3 functions or methods, at least one of which must return a value that is used
# somewhere else in your code. To clarify, at least one function should be called in your code, that function should
# calculate, retrieve, or otherwise set the value of a variable or data structure, return a value to where it was
# called, and use that value somewhere else in your code. For example, you could create a function that reads how many
# items there are in a text file, returns that value, and later uses that value to execute a loop a certain number
# of times.

# Create a class, then create at least one object of that class and populate it with data. The value of at least one
# object must be used somewhere in your code
